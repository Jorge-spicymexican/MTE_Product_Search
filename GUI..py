# Licensed to TCI, LLC
# Located in W132N10611 Germantown, Wisconsin
# This file is owned and copyrighted by TCI, LLC
#  Author: Jorge Jesus Jurado-Garcia
#  Title: Product Specialist Intern
#   Project Description: Marketing department MTE Product selection
#   Goal: create an easy to use GUI where the marketing department can
#          use for any price increases. Anyone with no background in CCS
#
#  Date of Creation: 12/29/2021
#  Rev:

import tkinter as tk
from tkinter import filedialog as fd
from tkinter import ttk
from tkinter.messagebox import showerror, showwarning, showinfo
from openpyxl import load_workbook
from selenium import webdriver
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException, ReadOnlyWorkbookException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


def file_lookup():
    filetypes = (
        ('All files', '*.*'),
        ('Application files', '*.exe'),
        ('Excel files', '*.xlsx')
    )
    filename = fd.askopenfilename(
        title='Insert File',
        initialdir='/',
        filetypes=filetypes)
    if filename != "":
        showinfo(
            title='Selected File',
            message=filename
        )
    return filename


def load_excel(file):
    null_array = []
    try:
        wb = load_workbook(file)
    except InvalidFileException:
        showerror(
            title='Error-Invalid File',
            message='Does not support old .xls file format. Please converter to a more recent excel format.'
        )
        return null_array
    except ReadOnlyWorkbookException:
        showerror(
            title='Error-Read Only Excel',
            message='Does not support Read only Excel Files, Select a different excel.'
        )
        return null_array
    if 'Sheet1' in wb.sheetnames:
        sh = wb["Sheet1"]
        row_ct = sh.max_row
        array_Product_Name = []
        for i in range(1, row_ct):
            array_Product_Name.append(sh.cell(row=i, column=1).value)
            # Will have to know how to grab information from excel and such
        return array_Product_Name
    else:
        showwarning(
            title='Warning- Sheet is not named 1',
            message='Make sure your Excel Sheet is name Sheet1'
        )
        return null_array


def selenium(array_product_name, web_driver, target, date, author, directory):
    # main website of MTE Corporation
    main_page = "https://www.mtecorp.com/click-find/"
    driver = webdriver.Chrome(executable_path=web_driver)
    # Goes to main page
    driver.get(main_page)
    # to maximize the browser window
    driver.minimize_window()
    driver = go(target, driver)
    wb = scrape(array_product_name, driver)
    save_excel(wb, date, author, directory)


def go(selection, driver):
    if selection == "RL":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[1]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "RLW":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[2]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
        # input_product_name(array_Product_Name)
    elif selection == "DVS":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[3]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "SWG":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[5]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "SWN":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[6]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "SWGM":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[7]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "MAP":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[8]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "MAEP":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[9]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "DVT":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[3]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    elif selection == "RF3":
        target = driver.find_element(By.XPATH, "/html/body/main/article/div/div/div/table/tbody/tr[12]/td[1]/a")
        driver.execute_script("arguments[0].click();", target)
    return driver


def scrape(array_product_name, driver):
    # these two arrays will be used to store the values havest from the webpage
    description = []
    unit_price = []
    #  switches where selenium looks at by going to the second screen
    #  function below is just so the bot can focus on new opened windows handler
    driver.switch_to.window(driver.window_handles[1])

    # for loop for the size of the array named this look will clear search bar, enter excel product name,
    # press enter, locate the next button, and click on the button for the next page.
    for ll in range(len(array_product_name)):
        # clears search bar
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".plxsty_pid"))).clear()
        # inputs text string into the search bar and waits to execute for 20 seconds
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".plxsty_pid"))).send_keys(
            array_product_name[ll])
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".plxsty_pid"))).send_keys(
            Keys.ENTER)
        # looks for element by full xpath and clicks with arguements[0] is are fullfilled.
        accept_bar = driver.find_element(By.XPATH, "/html/body/div[2]/div[1]/div[3]/table/tbody/tr/td[1]")
        driver.execute_script("arguments[0].click();", accept_bar)
        d1 = driver.find_element(By.XPATH, "/html/body/div/table[4]/tbody/tr[2]/td/table/tbody/tr[2]/td[3]")
        d1 = d1.text
        description.append(d1)
        p1 = driver.find_element(By.XPATH, "/html/body/div/table[4]/tbody/tr[2]/td/table/tbody/tr[2]/td[4]")
        p1 = p1.text
        unit_price.append(p1)
        # This line excute_script is to move back to the last opened page.
        driver.execute_script("window.history.go(-1)")

    # close the web browser and finish
    driver.close()
    wb = new_excel(array_product_name, description, unit_price)
    return wb


def new_excel(array_product_name, description, unit_price):
    # create a Workbook object.
    work_book = Workbook()
    sh = work_book.active
    sh.title = "Sheet1"
    for j in range(0, len(array_product_name)):
        sh.cell(row=j + 1, column=1).value = array_product_name[j]
        sh.cell(row=j + 1, column=2).value = description[j]
        sh.cell(row=j + 1, column=3).value = unit_price[j]
    return work_book


def save_excel(wb, date, author, directory):
    # create a file name using the date, and author
    # check if date has any / or \ -
    if directory != "":
        file = directory + "/" + author + "_" + date + "_" + "Result.xlsx"
    else:
        file = author + "_" + date + "_" + "Result.xlsx"
    wb.save(filename=file)
    showinfo(
        title='New Excel Location : ',
        message=file
    )


class MainFrame(ttk.Frame):
    # Initialization
    def __init__(self, container):
        super().__init__(container)
        # field options
        options = {'padx': 5, 'pady': 5}

        # Name label
        self.projectname_label = ttk.Label(self, text="User :")
        self.projectname_label.grid(column=0, row=0, sticky=tk.W, **options)

        # Name entry
        self.projectname = tk.StringVar()
        self.projectname_entry = ttk.Entry(self, textvariable=self.projectname)
        self.projectname_entry.grid(column=1, row=0, sticky=tk.EW, **options)
        self.projectname_entry.focus()

        # Date label
        self.Date_label = ttk.Label(self, text="Date : (Only _ )")
        self.Date_label.grid(column=0, row=1, sticky=tk.EW, **options)

        # Date entry
        self.Date = tk.StringVar()
        self.Date_entry = ttk.Entry(self, textvariable=self.Date)
        self.Date_entry.grid(column=1, row=1, sticky=tk.EW, **options)
        self.Date_entry.focus()

        # MTE product description IE LABEL
        self.lf = ttk.LabelFrame(text="Please select Product Line?")
        self.lf.grid(column=0, row=3, padx=20, pady=20)
        self.selected_product = tk.StringVar()
        selections = (('RL Reactor', 'RL'),
                      ('RLW Reactor', 'RLW'),
                      ('DV E-Series Filter', 'DVT'),
                      ('DV Sentry Filter', 'DVS'),
                      ('Sinewave Guardian', 'SWG'),
                      ('Sinewave Nexus', 'SWN'),
                      ('High Freq. Sinewave Gaurdian', 'SWGM'),
                      ('Matrix AP Filters', 'MAP'),
                      ('Matrix E-Series Filters', 'MAEP'),
                      ('RFI EMI Filters', 'RF3'),
                      )
        foof = True
        grid_row = 3
        for selection in selections:
            # create a radio button
            self.radio = ttk.Radiobutton(self.lf, text=selection[0], value=selection[1], variable=self.selected_product)
            if not foof:
                self.radio.grid(column=1, row=grid_row, ipadx=10, ipady=10)
                foof = True
            else:
                self.radio.grid(column=0, row=grid_row, ipadx=10, ipady=10)
                foof = False
                grid_row -= 1
            # grid column
            grid_row += 1

        # file_name
        self.filename = str()

        # web driver location
        self.web_driver = str()

        # file_name
        self.new_excel_path = str()
        # settings button
        self.button = ttk.Button(self, text="Finished Setting", command=self.settings)
        self.button.grid(row=0, column=3, sticky=tk.EW, **options)

        # Dump excel here
        self.button = ttk.Button(self, text="New Excel Location", command=self.new_excel_location)
        self.button.grid(row=9, column=3, sticky=tk.EW, **options)

        # excel button
        self.exel_button = ttk.Button(self, text='Insert Excel File', command=self.select_excel)
        self.exel_button.grid(column=0, row=9, sticky=tk.EW, **options)

        # web driver button
        self.web_button = ttk.Button(self, text='Select Web driver', command=self.select_webdriver)
        self.web_button.grid(column=1, row=9, sticky=tk.EW, **options)

        # add padding to the frame and show it
        self.grid(padx=10, pady=10, sticky=tk.NSEW)

    def new_excel_location(self):
        self.new_excel_path = self.folder_lookup()

    def settings(self):
        self.check_entry

    def show_selected_product(self):
        showinfo(
            title='Result',
            message=self.selected_product.get()
        )

    @staticmethod
    def folder_lookup():
        directory = fd.askdirectory(
        )
        if directory != "":
            showinfo(
                title='Selected Directory',
                message=directory
            )
        return directory

    @property
    def check_entry(self):
        check = True
        # Retrieve the of name, date, and MTE Selection
        inputs = [self.projectname.get(), self.Date.get(), self.selected_product.get(), self.filename,
                  self.web_driver]
        if inputs[0] == '':
            showerror(
                title='Error-Name',
                message='Please type in name.'
            )
            check = False
        if inputs[1] == '':
            showerror(
                title='Error-Date',
                message='Please type in Date.'
            )
            check = False
        if inputs[2] == '':
            showerror(
                title='Error-Selection',
                message='User did not Selected Product Line, Please Check.'
            )
            check = False
        if inputs[3] == '':
            showerror(
                title='Error-Excel File',
                message='User did not Inserted an Excel File.'
            )
            check = False
        if inputs[4] == '':
            showerror(
                title='Error-Webdriver',
                message='User did not Inserted a Selenium based Webdriver.'
            )
            check = False
        if check:
            showinfo(
                title='Settings',
                message='Settings are Configured'
            )
            self.show_selected_product()
            product_name = load_excel(inputs[3])
            if len(product_name) != 0:
                if self.web_driver[-16:] == "chromedriver.exe":
                    selenium(product_name, inputs[4], inputs[2], inputs[1], inputs[0], self.new_excel_path)
                else:
                    showerror(
                        title='Error-Driver',
                        message='Does not support this Driver. Make sure your Driver is named like this...chromedriver'
                    )

    def select_excel(self):
        self.filename = file_lookup()

    def select_webdriver(self):
        self.web_driver = file_lookup()


class App(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title('TCI LLC - MTE Product Search')
        self.resizable(False, False)
        # ensure that a window is always at the top of the stacking order
        self.attributes('-topmost', 1)

        window_width = 400
        window_height = 400

        # get screen dimension
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()

        # find the center point
        center_x = int(screen_width / 2 - window_width / 2)
        center_y = int(screen_height / 2 - window_height / 2)

        # create the screen on window console
        self.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')

        # changing the Tinker Logo into the TCI logo instead for our development
        self.iconbitmap('./assets/tci_logo_Csx_icon.ico')
        frm = ttk.Frame(self, padding=1)
        frm.grid()


if __name__ == "__main__":
    app = App()
    MainFrame(app)
    app.mainloop()
